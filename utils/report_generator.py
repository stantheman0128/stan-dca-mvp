# utils/report_generator.py
"""Report generation module for Excel and PDF exports."""

import io
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional

import pandas as pd

# Excel
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# PDF
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib import colors
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


class ReportGenerator:
    """
    Generates Excel and PDF reports for backtest results.
    """
    
    @staticmethod
    def export_to_excel(
        results: List[Any],  # List[BacktestResult]
        comparison_df: Optional[pd.DataFrame] = None,
        output_path: Optional[str] = None
    ) -> bytes:
        """
        Export backtest results to Excel file.
        
        Args:
            results: List of BacktestResult objects
            comparison_df: Optional comparison DataFrame
            output_path: Optional file path to save (if None, returns bytes)
            
        Returns:
            Excel file as bytes
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export")
        
        wb = Workbook()
        
        # Styles
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2196F3', end_color='2196F3', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Sheet 1: Summary
        ws_summary = wb.active
        ws_summary.title = '概述'
        
        # Title
        ws_summary['A1'] = 'DCA 策略回測結果報告'
        ws_summary['A1'].font = Font(bold=True, size=16)
        ws_summary['A2'] = f'生成時間: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws_summary.append([])
        
        # Key metrics comparison
        if results:
            ws_summary.append(['策略比較摘要'])
            ws_summary.append([])
            
            headers = ['策略', '總報酬率', '年化報酬率', '最大回撤', '夏普比率', '總投入', '最終市值']
            ws_summary.append(headers)
            
            for cell in ws_summary[ws_summary.max_row]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
            
            for r in results:
                m = r.metrics
                ws_summary.append([
                    r.strategy_name,
                    f'{m.total_return:.2f}%',
                    f'{m.cagr:.2f}%',
                    f'{m.max_drawdown:.2f}%',
                    f'{m.sharpe_ratio:.2f}',
                    f'{m.total_invested:,.0f}',
                    f'{m.final_value:,.0f}',
                ])
        
        # Sheet 2: Detailed Metrics
        ws_metrics = wb.create_sheet('詳細指標')
        
        if comparison_df is not None and not comparison_df.empty:
            for r_idx, row in enumerate(dataframe_to_rows(comparison_df, index=True, header=True)):
                ws_metrics.append(row)
                if r_idx == 0:
                    for cell in ws_metrics[ws_metrics.max_row]:
                        cell.font = header_font
                        cell.fill = header_fill
        
        # Sheet 3+: Individual strategy transactions
        for result in results:
            safe_name = result.strategy_name[:20].replace('/', '_')
            ws_trans = wb.create_sheet(f'{safe_name}_交易')
            
            # Add transaction data
            trans_df = result.transactions[['date', 'price', 'investment', 'shares_bought', 
                                            'total_shares', 'total_cost', 'current_value', 'return_pct']]
            trans_df = trans_df.copy()
            trans_df['date'] = trans_df['date'].dt.strftime('%Y-%m-%d')
            trans_df.columns = ['日期', '價格', '投入金額', '買入股數', '累積股數', '累積成本', '市值', '報酬率(%)']
            
            for r_idx, row in enumerate(dataframe_to_rows(trans_df, index=False, header=True)):
                ws_trans.append(row)
                if r_idx == 0:
                    for cell in ws_trans[ws_trans.max_row]:
                        cell.font = header_font
                        cell.fill = header_fill
        
        # Adjust column widths
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        if output_path:
            with open(output_path, 'wb') as f:
                f.write(output.getvalue())
        
        return output.getvalue()
    
    @staticmethod
    def generate_pdf_report(
        results: List[Any],
        title: str = "DCA 策略回測研究報告",
        output_path: Optional[str] = None
    ) -> bytes:
        """
        Generate a PDF report for backtest results.
        
        Args:
            results: List of BacktestResult objects
            title: Report title
            output_path: Optional file path to save
            
        Returns:
            PDF file as bytes
        """
        if not REPORTLAB_AVAILABLE:
            raise ImportError("reportlab is required for PDF export")
        
        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=A4,
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=0.5*inch,
            bottomMargin=0.5*inch
        )
        
        # Styles (using default fonts for compatibility)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=20,
            alignment=1  # Center
        )
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=10,
            textColor=colors.HexColor('#2196F3')
        )
        normal_style = styles['Normal']
        
        elements = []
        
        # Title
        elements.append(Paragraph(title, title_style))
        elements.append(Paragraph(
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            normal_style
        ))
        elements.append(Spacer(1, 20))
        
        # Executive Summary
        elements.append(Paragraph("Executive Summary", heading_style))
        
        if results:
            # Find best strategy by Sharpe ratio
            best_sharpe = max(results, key=lambda r: r.metrics.sharpe_ratio)
            best_return = max(results, key=lambda r: r.metrics.total_return)
            lowest_dd = min(results, key=lambda r: r.metrics.max_drawdown)
            
            summary_text = f"""
            Test Period: {results[0].start_date} to {results[0].end_date}<br/>
            Number of Strategies Tested: {len(results)}<br/>
            Best Sharpe Ratio: {best_sharpe.strategy_name} ({best_sharpe.metrics.sharpe_ratio:.2f})<br/>
            Highest Return: {best_return.strategy_name} ({best_return.metrics.total_return:.1f}%)<br/>
            Lowest Drawdown: {lowest_dd.strategy_name} ({lowest_dd.metrics.max_drawdown:.1f}%)
            """
            elements.append(Paragraph(summary_text, normal_style))
            elements.append(Spacer(1, 15))
        
        # Strategy Comparison Table
        elements.append(Paragraph("Strategy Comparison", heading_style))
        
        if results:
            table_data = [['Strategy', 'Total Return', 'CAGR', 'Max DD', 'Sharpe', 'Sortino']]
            
            for r in results:
                m = r.metrics
                table_data.append([
                    r.strategy_name[:15],
                    f'{m.total_return:.1f}%',
                    f'{m.cagr:.1f}%',
                    f'{m.max_drawdown:.1f}%',
                    f'{m.sharpe_ratio:.2f}',
                    f'{m.sortino_ratio:.2f}',
                ])
            
            table = Table(table_data, colWidths=[1.8*inch, 1*inch, 0.8*inch, 0.8*inch, 0.7*inch, 0.7*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2196F3')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(table)
            elements.append(Spacer(1, 20))
        
        # Individual Strategy Details
        elements.append(Paragraph("Strategy Details", heading_style))
        
        for r in results:
            m = r.metrics
            detail_text = f"""
            <b>{r.strategy_name}</b><br/>
            - Total Invested: {m.total_invested:,.0f}<br/>
            - Final Value: {m.final_value:,.0f}<br/>
            - Total Return: {m.total_return:.2f}%<br/>
            - Annualized Return (CAGR): {m.cagr:.2f}%<br/>
            - Maximum Drawdown: {m.max_drawdown:.2f}%<br/>
            - Annual Volatility: {m.volatility:.2f}%<br/>
            - Sharpe Ratio: {m.sharpe_ratio:.2f}<br/>
            - Sortino Ratio: {m.sortino_ratio:.2f}<br/>
            - Calmar Ratio: {m.calmar_ratio:.2f}<br/>
            - Number of Trades: {m.total_trades}<br/>
            - Win Rate: {m.win_rate:.1f}%
            """
            elements.append(Paragraph(detail_text, normal_style))
            elements.append(Spacer(1, 10))
        
        # Disclaimer
        elements.append(Spacer(1, 30))
        elements.append(Paragraph("Disclaimer", heading_style))
        disclaimer = """
        This report is for informational purposes only. Past performance is not 
        indicative of future results. Investment involves risk, and you may lose 
        some or all of your investment. Always do your own research before making 
        investment decisions.
        """
        elements.append(Paragraph(disclaimer, normal_style))
        
        # Build PDF
        doc.build(elements)
        output.seek(0)
        
        if output_path:
            with open(output_path, 'wb') as f:
                f.write(output.getvalue())
        
        return output.getvalue()
    
    @staticmethod
    def export_transactions_csv(
        result: Any,  # BacktestResult
        output_path: Optional[str] = None
    ) -> str:
        """
        Export transaction history to CSV.
        
        Args:
            result: BacktestResult object
            output_path: Optional file path
            
        Returns:
            CSV content as string
        """
        df = result.transactions.copy()
        df['date'] = df['date'].dt.strftime('%Y-%m-%d')
        
        csv_content = df.to_csv(index=False)
        
        if output_path:
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(csv_content)
        
        return csv_content
